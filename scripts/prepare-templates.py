"""Maintain the checked-in Excel templates without changing their business headers.

Optional regeneration dependency: openpyxl. Runtime application does not need Python.
"""
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation


def prepare(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook["导入数据"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, sheet.max_column).column_letter}1001"
    for validation in sheet.data_validations.dataValidation:
        ranges = []
        for existing in validation.sqref.ranges:
            target = CellRange(str(existing))
            target.max_row = 1001
            ranges.append(str(target))
        validation.sqref = " ".join(ranges)
        validation.allow_blank = True
        validation.showErrorMessage = True
        validation.errorStyle = "stop"
        validation.errorTitle = "填写内容无效"
        validation.error = "请使用下拉选项填写，上传时还将进行服务端校验。"

    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(2, column)
        header = sheet.cell(1, column).value
        for row in range(3, 1002):
            cell = sheet.cell(row, column)
            cell._style = copy(source._style)
        target_range = f"{source.column_letter}2:{source.column_letter}1001"
        validation = None
        if header in ("合同总金额（含税）", "节点金额", "实收金额"):
            validation = DataValidation(type="decimal", operator="greaterThan", formula1="0")
            for row in range(2, 1002):
                sheet.cell(row, column).number_format = '#,##0.00'
        elif header in ("节点序号", "账期天数"):
            validation = DataValidation(type="whole", operator="between",
                formula1="0" if header == "账期天数" else "1",
                formula2="3650" if header == "账期天数" else "2147483647")
        elif header in ("合同签订日期", "基准日期", "实收日期"):
            validation = DataValidation(type="date", operator="between", formula1='DATE(1900,1,1)', formula2='DATE(9999,12,31)')
            for row in range(2, 1002):
                sheet.cell(row, column).number_format = 'yyyy-mm-dd'
        if validation is not None:
            # Replace only the same generated numeric/date validator on repeated runs.
            sheet.data_validations.dataValidation = [item for item in sheet.data_validations.dataValidation
                if not (item.type == validation.type and str(item.sqref) == target_range)]
            validation.allow_blank = True
            validation.showErrorMessage = True
            validation.errorStyle = "stop"
            validation.errorTitle = "请检查输入"
            validation.error = "日期使用 YYYY-MM-DD，金额大于0且最多两位小数，序号与账期使用整数。"
            validation.add(target_range)
            sheet.add_data_validation(validation)
    workbook.save(path)
    print(f"已校验模板：{path.name}（冻结表头，1000行输入格式与校验）")


if __name__ == "__main__":
    for template in sorted((Path(__file__).resolve().parent.parent / "public" / "templates").glob("*.xlsx")):
        prepare(template)
